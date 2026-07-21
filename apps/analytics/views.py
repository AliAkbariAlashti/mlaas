import csv
from pathlib import Path

import xlrd
from django.db import transaction
from django.db.models import Count
from django.shortcuts import get_object_or_404
from drf_spectacular.utils import extend_schema
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.generics import GenericAPIView, ListAPIView, RetrieveAPIView
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from .models import AnalysisService, Dataset, Project, RunEvent, WaitlistLead
from .serializers import (
    BasketResultSerializer,
    DashboardSerializer,
    DatasetCreateSerializer,
    DatasetSerializer,
    MappingSerializer,
    PredictiveResultSerializer,
    ProjectStatusSerializer,
    ProjectHistorySerializer,
    ProjectUploadResponseSerializer,
    ProjectUploadSerializer,
    RFMResultSerializer,
    RunEventSerializer,
    ServiceSerializer,
    StartAnalysisResponseSerializer,
    WaitlistResponseSerializer,
)
from .tasks import run_analysis_task
from apps.developer_api.models import APIKey


def extract_headers(uploaded_file) -> list[str]:
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        uploaded_file.seek(0)
        line = uploaded_file.readline().decode("utf-8-sig")
        headers = next(csv.reader([line]))
    elif suffix == ".xlsx":
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        headers = next(workbook.active.iter_rows(values_only=True), ())
        workbook.close()
    else:
        uploaded_file.seek(0)
        workbook = xlrd.open_workbook(file_contents=uploaded_file.read(), on_demand=True)
        sheet = workbook.sheet_by_index(0)
        headers = sheet.row_values(0) if sheet.nrows else []
        workbook.release_resources()
    uploaded_file.seek(0)
    cleaned = [str(value).strip() for value in headers if value is not None and str(value).strip()]
    if not cleaned:
        raise ValueError("The uploaded file has no header row.")
    return cleaned


class ServiceListView(ListAPIView):
    queryset = AnalysisService.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = (AllowAny,)


class ProjectHistoryView(ListAPIView):
    serializer_class = ProjectHistorySerializer

    def get_queryset(self):
        return Project.objects.filter(user=self.request.user).select_related("service", "dataset")


class DatasetListCreateView(GenericAPIView):
    serializer_class = DatasetSerializer

    def get(self, request):
        datasets = Dataset.objects.filter(user=request.user).annotate(runs_count=Count("runs"))
        return Response(DatasetSerializer(datasets, many=True).data)

    @extend_schema(request=DatasetCreateSerializer, responses={201: DatasetSerializer})
    def post(self, request):
        serializer = DatasetCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        uploaded_file = serializer.validated_data["file"]
        try:
            detected_columns = extract_headers(uploaded_file)
        except (UnicodeDecodeError, csv.Error, ValueError, xlrd.XLRDError) as exc:
            return Response({"file": [str(exc)]}, status=status.HTTP_400_BAD_REQUEST)
        dataset = Dataset.objects.create(
            user=request.user,
            name=serializer.validated_data["name"],
            file=uploaded_file,
            original_filename=uploaded_file.name,
            file_type=Path(uploaded_file.name).suffix.lower().lstrip("."),
            file_size=uploaded_file.size,
            detected_columns=detected_columns,
            validation_status=Dataset.ValidationStatus.VALID,
        )
        dataset.runs_count = 0
        return Response(DatasetSerializer(dataset).data, status=status.HTTP_201_CREATED)


class DatasetDetailView(RetrieveAPIView):
    serializer_class = DatasetSerializer
    lookup_url_kwarg = "dataset_id"

    def get_queryset(self):
        return Dataset.objects.filter(user=self.request.user).annotate(runs_count=Count("runs"))


class DashboardView(GenericAPIView):
    serializer_class = DashboardSerializer

    def get(self, request):
        projects = Project.objects.filter(user=request.user).select_related("service", "dataset")
        recent = projects[:5]
        return Response({
            "credits_remaining": request.user.credit_limit,
            "total_projects": projects.count(),
            "successful_projects": projects.filter(status=Project.Status.SUCCESS).count(),
            "processing_projects": projects.filter(status=Project.Status.PROCESSING).count(),
            "failed_projects": projects.filter(status=Project.Status.FAILED).count(),
            "waitlist_requests": WaitlistLead.objects.filter(user=request.user).count(),
            "recent_projects": ProjectHistorySerializer(recent, many=True).data,
        })


class ProjectUploadView(GenericAPIView):
    serializer_class = ProjectUploadSerializer

    @extend_schema(responses={201: ProjectUploadResponseSerializer})
    def post(self, request):
        serializer = ProjectUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        if isinstance(request.auth, APIKey):
            subscription = request.user.api_subscription
            service = serializer.validated_data["analysis_type"]
            if not subscription.plan.services.filter(pk=service.pk).exists():
                return Response(
                    {"detail": "This analytics service is not included in the API plan."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        try:
            detected_columns = extract_headers(serializer.validated_data["file"])
        except (UnicodeDecodeError, csv.Error, ValueError, xlrd.XLRDError) as exc:
            return Response({"file": [str(exc)]}, status=status.HTTP_400_BAD_REQUEST)
        serializer.context["detected_columns"] = detected_columns
        project = serializer.save(user=request.user)
        return Response({
            "project_id": str(project.id),
            "analysis_type": project.analysis_type,
            "detected_columns": detected_columns,
        }, status=status.HTTP_201_CREATED)


class OwnedProjectView(GenericAPIView):
    def get_project(self, project_id):
        return get_object_or_404(Project.objects.select_related("service"), pk=project_id, user=self.request.user)


class ResumeProjectView(OwnedProjectView):
    serializer_class = ProjectUploadResponseSerializer

    def get(self, request, project_id):
        project = self.get_project(project_id)
        if project.status != Project.Status.PENDING:
            return Response(
                {"detail": "Only pending projects can be resumed."},
                status=status.HTTP_409_CONFLICT,
            )
        if not project.raw_file_path:
            return Response(
                {"detail": "The uploaded source file is no longer available."},
                status=status.HTTP_410_GONE,
            )
        try:
            with project.raw_file_path.open("rb") as uploaded_file:
                detected_columns = extract_headers(uploaded_file)
        except (FileNotFoundError, UnicodeDecodeError, csv.Error, ValueError, xlrd.XLRDError) as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({
            "project_id": str(project.id),
            "analysis_type": project.analysis_type,
            "detected_columns": detected_columns,
        })


class StartAnalysisView(OwnedProjectView):
    serializer_class = MappingSerializer

    @extend_schema(responses={202: StartAnalysisResponseSerializer})
    def post(self, request, project_id):
        project = self.get_project(project_id)
        if not project.service.is_active:
            return Response({"detail": "This service is available through the private beta waitlist."}, status=status.HTTP_409_CONFLICT)
        if project.status != Project.Status.PENDING:
            return Response({"detail": "This project has already been submitted."}, status=status.HTTP_409_CONFLICT)
        serializer = MappingSerializer(data=request.data, context={"project": project})
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            user = type(request.user).objects.select_for_update().get(pk=request.user.pk)
            if user.credit_limit < 1:
                return Response({"detail": "No analysis credits remain."}, status=status.HTTP_402_PAYMENT_REQUIRED)
            user.credit_limit -= 1
            user.save(update_fields=("credit_limit",))
            project.data_mapping = serializer.validated_data["mapping"]
            project.status = Project.Status.PROCESSING
            project.save(update_fields=("data_mapping", "status"))
            RunEvent.objects.create(run=project, stage=RunEvent.Stage.MAPPING, message="Input schema mapping validated.")
            RunEvent.objects.create(run=project, stage=RunEvent.Stage.QUEUED, message="Run submitted to the execution queue.")
            transaction.on_commit(lambda: run_analysis_task.delay(str(project.id)))
        return Response({
            "project_id": str(project.id),
            "status": Project.Status.PROCESSING,
            "message": "Data ingestion complete. Analysis added to Celery queue.",
        }, status=status.HTTP_202_ACCEPTED)


class ProjectStatusView(OwnedProjectView):
    serializer_class = ProjectStatusSerializer

    def get(self, request, project_id):
        project = self.get_project(project_id)
        response = {"status": project.status}
        if project.status == Project.Status.FAILED:
            response["error"] = project.error_log
        return Response(response)


class RunEventListView(OwnedProjectView):
    serializer_class = RunEventSerializer

    def get(self, request, project_id):
        project = self.get_project(project_id)
        return Response(RunEventSerializer(project.events.all(), many=True).data)


class RFMResultView(OwnedProjectView):
    serializer_class = RFMResultSerializer

    def get(self, request, project_id):
        result = self.get_project(project_id).rfm_result
        return Response({
            "summary": result.summary,
            "chart_data": result.chart_data,
            "actionable_insights": result.actionable_insights,
            "download_excel_url": result.result_file_path,
        })


class BasketResultView(OwnedProjectView):
    serializer_class = BasketResultSerializer

    def get(self, request, project_id):
        rules = self.get_project(project_id).basket_result.rules
        confidence_values = [rule.get("confidence", 0) for rule in rules]
        lift_values = [rule.get("lift", 0) for rule in rules]
        return Response({
            "summary": {
                "rules_found": len(rules),
                "strongest_lift": round(max(lift_values), 4) if lift_values else 0,
                "average_confidence_percentage": round(
                    sum(confidence_values) * 100 / len(confidence_values), 2
                ) if confidence_values else 0,
            },
            "rules": rules,
        })


class PredictiveResultView(OwnedProjectView):
    serializer_class = PredictiveResultSerializer

    def get(self, request, project_id):
        project = self.get_project(project_id)
        result = project.ml_result
        payload = {"analysis_type": project.analysis_type, "summary": result.metrics}
        payload["anomalies_list" if project.analysis_type == "ANOMALY" else "scores"] = result.visualization_data
        return Response(payload)


class JoinWaitlistView(OwnedProjectView):
    serializer_class = WaitlistResponseSerializer

    @extend_schema(request=None)
    def post(self, request, project_id):
        project = self.get_project(project_id)
        if project.service.is_active:
            return Response(
                {"detail": "This service is already available and does not accept beta requests."},
                status=status.HTTP_409_CONFLICT,
            )
        if project.status == Project.Status.WAITLISTED:
            return Response({
                "status": "waitlisted",
                "message": "Your private-beta request is already registered.",
            })
        if project.status != Project.Status.PENDING:
            return Response(
                {"detail": "Only pending projects can join the private beta."},
                status=status.HTTP_409_CONFLICT,
            )
        WaitlistLead.objects.get_or_create(project=project, defaults={"user": request.user})
        project.status = Project.Status.WAITLISTED
        project.save(update_fields=("status",))
        return Response({
            "status": "waitlisted",
            "message": "Your request for this automated engine has been registered. Our accounts team will contact you shortly to run an isolated private-beta trial on your raw data.",
        })
